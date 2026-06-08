from pathlib import Path
from collections import Counter
from html import escape
from flask import Flask, render_template_string
from openpyxl import load_workbook

app = Flask(__name__)

HTML = """<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Donovan Ticket Dashboard</title>
<style>
body{font-family:Arial,Helvetica,sans-serif;background:#f3f6fa;margin:0;padding:24px;color:#111}h1{color:#083b70;margin:0 0 20px}h2{margin-top:28px}.cards{display:flex;gap:20px;flex-wrap:wrap}.card{background:white;border-radius:10px;box-shadow:0 3px 12px #d7dde5;padding:22px 26px;min-width:210px}.card .label{font-weight:700;font-size:18px}.card .num{font-size:26px;font-weight:800;margin-top:20px}table{border-collapse:collapse;background:white;width:100%;margin-top:12px}th{background:#063d72;color:white;text-align:left;padding:10px}td{border:1px solid #ddd;padding:9px;vertical-align:top}tr:nth-child(even) td{background:#fbfbfb}.small{color:#555;margin-top:-10px}.stuck{font-weight:700;color:#8a1f11}@media(max-width:760px){body{padding:14px}.card{min-width:150px}table{font-size:12px}}
</style></head><body>
<h1>Donovan Ticket Dashboard</h1>
<div class="small">Updated from {{ filename }}</div>
<div class="cards">
 <div class="card"><div class="label">Total Tickets</div><div class="num">{{ total }}</div></div>
 <div class="card"><div class="label">Active Tickets</div><div class="num">{{ active }}</div></div>
 <div class="card"><div class="label">Tickets Stuck &gt;3 Days</div><div class="num stuck">{{ stuck }}</div></div>
 <div class="card"><div class="label">Average Days In Status</div><div class="num">{{ avg_days }}</div></div>
</div>
<h2>Status Summary</h2>
<table><thead><tr><th>Status</th><th style="width:120px">Count</th></tr></thead><tbody>
{% for status, count in status_counts %}<tr><td>{{ status }}</td><td>{{ count }}</td></tr>{% endfor %}
</tbody></table>
<h2>Ticket Details</h2>
<table><thead><tr><th>Job Number</th><th>Status</th><th>Days In Status</th><th>Assigned To</th><th>Customer</th><th>Location</th><th>Description</th></tr></thead><tbody>
{% for r in rows %}<tr><td>{{ r.get('Job Number','') }}</td><td>{{ r.get('Job Status','') }}</td><td>{{ r.get('Days In Status','') }}</td><td>{{ r.get('Assigned To','') }}</td><td>{{ r.get('Customer','') }}</td><td>{{ r.get('Location','') }}</td><td>{{ r.get('Description','') }}</td></tr>{% endfor %}
</tbody></table>
</body></html>"""

def _truthy(value):
    return str(value).strip().lower() in {"true", "1", "yes", "y"}

def _closed(value):
    return str(value).strip().lower() in {"true", "1", "yes", "closed"}

def _num(value):
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except Exception:
        return 0.0

def find_latest_excel():
    files = sorted(Path('.').glob('Job Card List*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None

def read_excel_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    records = []
    for row in rows[1:]:
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        records.append(rec)
    return records

def build_records():
    latest = find_latest_excel()
    if not latest:
        return [], "No Job Card List Excel file found"

    records = read_excel_rows(latest)
    filtered = []
    for r in records:
        assigned = str(r.get('Assigned To', '') or '')
        if 'Donovan Willemse' not in assigned:
            continue
        if 'Is Active' in r and not _truthy(r.get('Is Active')):
            continue
        if 'Closed' in r and _closed(r.get('Closed')):
            continue
        filtered.append(r)
    return filtered, latest.name

@app.route('/')
def dashboard():
    rows, filename = build_records()
    days = [_num(r.get('Days In Status')) for r in rows]
    total = len(rows)
    stuck = sum(1 for d in days if d > 3)
    avg_days = round(sum(days) / len(days), 1) if days else 0
    status_counts = Counter(str(r.get('Job Status', '') or '') for r in rows).most_common()
    return render_template_string(
        HTML,
        filename=filename,
        rows=rows,
        total=total,
        active=total,
        stuck=stuck,
        avg_days=avg_days,
        status_counts=status_counts,
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
